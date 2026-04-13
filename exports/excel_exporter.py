"""
Excel Exporter - Genera file Excel multi-sheet per analisi Ericsson
9 fogli: RadioUnits, Allarmi, FRU, PUSCH, RET, TMA, Fiber, SFP, BranchPairs
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


# Colori intestazioni per ogni sheet
COLOR_BLUE   = "4472C4"
COLOR_RED    = "C00000"
COLOR_ORANGE = "ED7D31"
COLOR_GREEN  = "70AD47"
COLOR_PURPLE = "7030A0"
COLOR_TEAL   = "00B0F0"
COLOR_GREY   = "595959"
COLOR_GOLD   = "BF8F00"
COLOR_DARK   = "243F60"
COLOR_NAVY   = "1A3A5C"

# Colori evidenziazione valori critici
BG_CRITICAL  = "FFCCCC"  # rosso chiaro
BG_WARNING   = "FFF2CC"  # giallo chiaro
BG_NORMAL    = "E2EFDA"  # verde chiaro


class ExcelExporter:
    """Genera Excel professionale con 9 sheet formattati"""

    def __init__(self, analysis):
        self.analysis = analysis
        self.wb = Workbook()
        self.wb.remove(self.wb.active)   # rimuove sheet vuoto default

    def generate(self) -> BytesIO:
        """Genera il workbook e restituisce BytesIO"""
        self._sheet_radio_units()
        self._sheet_allarmi()
        self._sheet_fru()
        self._sheet_pusch()
        self._sheet_ret()
        self._sheet_tma()
        self._sheet_fiber()
        self._sheet_sfp()
        self._sheet_branch_pairs()
        self._sheet_lga()

        excel_file = BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

    # ------------------------------------------------------------------ #
    #  Utility                                                             #
    # ------------------------------------------------------------------ #

    def _header_style(self, ws, headers, color):
        """Applica stile intestazione con colore scelto"""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _autosize(self, ws):
        """Autosize colonne in base al contenuto"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def _row_fill(self, ws, row_idx, color):
        """Colora una riga intera"""
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # ------------------------------------------------------------------ #
    #  1. Radio Units VSWR                                                 #
    # ------------------------------------------------------------------ #

    def _sheet_radio_units(self):
        ws = self.wb.create_sheet("Radio Units VSWR")
        headers = ['FRU', 'Board', 'RF Port', 'Branch Pair', 'TX', 'TX Unit',
                   'VSWR', 'Return Loss (dB)', 'RX (dBm)', 'VSWR Warning', 'VSWR Critical']
        self._header_style(ws, headers, COLOR_BLUE)

        for idx, unit in enumerate(self.analysis.radio_units.all(), start=2):
            row = [
                unit.fru, unit.board, unit.rf_port, unit.branch_pair,
                float(unit.tx) if unit.tx else '',
                unit.tx_unit,
                float(unit.vswr),
                float(unit.return_loss) if unit.return_loss else '',
                float(unit.rx) if unit.rx else '',
                'Sì' if unit.is_vswr_warning else 'No',
                'Sì' if unit.is_vswr_critical else 'No',
            ]
            for col, val in enumerate(row, start=1):
                ws.cell(row=idx, column=col, value=val)

            # Evidenzia righe critiche
            if unit.is_vswr_critical:
                self._row_fill(ws, idx, BG_CRITICAL)
            elif unit.is_vswr_warning:
                self._row_fill(ws, idx, BG_WARNING)

        self._autosize(ws)

    # ------------------------------------------------------------------ #
    #  2. Allarmi                                                          #
    # ------------------------------------------------------------------ #

    def _sheet_allarmi(self):
        ws = self.wb.create_sheet("Allarmi")
        headers = ['Severity', 'Alarm Number', 'Causa']
        self._header_style(ws, headers, COLOR_RED)

        SEVERITY_COLORS = {
            'CRITICAL': BG_CRITICAL,
            'MAJOR': BG_WARNING,
            'MINOR': BG_NORMAL,
        }

        for idx, alarm in enumerate(self.analysis.alarms.all(), start=2):
            ws.cell(row=idx, column=1, value=alarm.severity)
            ws.cell(row=idx, column=2, value=alarm.alarm_number)
            ws.cell(row=idx, column=3, value=alarm.cause)
            color = SEVERITY_COLORS.get(alarm.severity, '')
            if color:
                self._row_fill(ws, idx, color)

        self._autosize(ws)

    # ------------------------------------------------------------------ #
    #  3. FRU                                                              #
    # ------------------------------------------------------------------ #

    def _sheet_fru(self):
        ws = self.wb.create_sheet("FRU")
        headers = ['Nome', 'Board', 'PM Temp', 'Temperatura (°C)', 'Temp Elevata (>60°C)']
        self._header_style(ws, headers, COLOR_ORANGE)

        for idx, fru in enumerate(self.analysis.fru_units.all(), start=2):
            ws.cell(row=idx, column=1, value=fru.name)
            ws.cell(row=idx, column=2, value=fru.board)
            ws.cell(row=idx, column=3, value=fru.pmtemp)
            ws.cell(row=idx, column=4, value=fru.temp if fru.temp is not None else 'N/A')
            ws.cell(row=idx, column=5, value='Sì' if fru.is_temp_high else 'No')
            if fru.is_temp_high:
                self._row_fill(ws, idx, BG_CRITICAL)

        self._autosize(ws)

    # ------------------------------------------------------------------ #
    #  4. PUSCH / PUCCH                                                    #
    # ------------------------------------------------------------------ #

    def _sheet_pusch(self):
        ws = self.wb.create_sheet("PUSCH PUCCH RSSI")
        headers = ['Cell', 'SC', 'FRU', 'Board', 'PUSCH', 'PUCCH',
                   'Port A', 'Port B', 'Port C', 'Port D', 'Delta', 'RSSI Alto']
        self._header_style(ws, headers, COLOR_PURPLE)

        for idx, row_data in enumerate(self.analysis.pusch_data.all(), start=2):
            row = [
                row_data.cell, row_data.sc, row_data.fru, row_data.board,
                float(row_data.pusch), float(row_data.pucch),
                float(row_data.port_a) if row_data.port_a else '',
                float(row_data.port_b) if row_data.port_b else '',
                float(row_data.port_c) if row_data.port_c else '',
                float(row_data.port_d) if row_data.port_d else '',
                float(row_data.delta) if row_data.delta else '',
                'Sì' if row_data.is_rssi_high else 'No',
            ]
            for col, val in enumerate(row, start=1):
                ws.cell(row=idx, column=col, value=val)
            if row_data.is_rssi_high:
                self._row_fill(ws, idx, BG_WARNING)

        self._autosize(ws)

    # ------------------------------------------------------------------ #
    #  5. RET Devices                                                      #
    # ------------------------------------------------------------------ #

    def _sheet_ret(self):
        ws = self.wb.create_sheet("RET Devices")
        headers = ['Antenna Group', 'Antenna Near Unit', 'Radio Unit',
                   'Status', 'Device Type', 'Product Nr', 'Revision', 'Unique ID']
        self._header_style(ws, headers, COLOR_TEAL)

        for idx, ret in enumerate(self.analysis.ret_devices.all(), start=2):
            ws.cell(row=idx, column=1, value=ret.antenna_group)
            ws.cell(row=idx, column=2, value=ret.antenna_near_unit)
            ws.cell(row=idx, column=3, value=ret.radio_unit)
            ws.cell(row=idx, column=4, value=ret.status)
            ws.cell(row=idx, column=5, value=ret.device_type)
            ws.cell(row=idx, column=6, value=ret.product_nr)
            ws.cell(row=idx, column=7, value=ret.revision)
            ws.cell(row=idx, column=8, value=ret.unique_id)

        self._autosize(ws)

    # ------------------------------------------------------------------ #
    #  6. TMA Devices                                                      #
    # ------------------------------------------------------------------ #

    def _sheet_tma(self):
        ws = self.wb.create_sheet("TMA Devices")
        headers = ['Antenna Group', 'Antenna Near Unit', 'Radio Unit',
                   'Status', 'Device Type', 'Product Nr', 'Revision', 'Unique ID']
        self._header_style(ws, headers, COLOR_GREEN)

        for idx, tma in enumerate(self.analysis.tma_devices.all(), start=2):
            ws.cell(row=idx, column=1, value=tma.antenna_group)
            ws.cell(row=idx, column=2, value=tma.antenna_near_unit)
            ws.cell(row=idx, column=3, value=tma.radio_unit)
            ws.cell(row=idx, column=4, value=tma.status)
            ws.cell(row=idx, column=5, value=tma.device_type)
            ws.cell(row=idx, column=6, value=tma.product_nr)
            ws.cell(row=idx, column=7, value=tma.revision)
            ws.cell(row=idx, column=8, value=tma.unique_id)

        self._autosize(ws)

    # ------------------------------------------------------------------ #
    #  7. Fiber Links                                                       #
    # ------------------------------------------------------------------ #

    def _sheet_fiber(self):
        ws = self.wb.create_sheet("Fiber Links")
        headers = ['Link ID', 'Status', 'RiL', 
                   'WL1 (nm)', 'Temp1 (°C)', 'TXbs1 (%)', 'TX1 (dBm)', 'RX1 (dBm)',
                   'WL2 (nm)', 'Temp2 (°C)', 'TXbs2 (%)', 'TX2 (dBm)', 'RX2 (dBm)',
                   'DL Loss (dB)', 'UL Loss (dB)', 'Length', 
                   'DL Critico', 'UL Critico', 'Link Down']
        self._header_style(ws, headers, COLOR_GREY)
        
        for idx, fiber in enumerate(self.analysis.fiber_links.all(), start=2):
            ws.cell(row=idx, column=1, value=fiber.link_id or '')
            ws.cell(row=idx, column=2, value=fiber.link_status or '')
            ws.cell(row=idx, column=3, value=fiber.ril or '')
            ws.cell(row=idx, column=4, value=float(fiber.wl1) if fiber.wl1 else '')
            ws.cell(row=idx, column=5, value=fiber.temp1 if fiber.temp1 else '')
            ws.cell(row=idx, column=6, value=fiber.txbs1 if fiber.txbs1 else '')
            ws.cell(row=idx, column=7, value=float(fiber.txdbm1) if fiber.txdbm1 else '')
            ws.cell(row=idx, column=8, value=float(fiber.rxdbm1) if fiber.rxdbm1 else '')
            ws.cell(row=idx, column=9, value=float(fiber.wl2) if fiber.wl2 else '')
            ws.cell(row=idx, column=10, value=fiber.temp2 if fiber.temp2 else '')
            ws.cell(row=idx, column=11, value=fiber.txbs2 if fiber.txbs2 else '')
            ws.cell(row=idx, column=12, value=float(fiber.txdbm2) if fiber.txdbm2 else '')
            ws.cell(row=idx, column=13, value=float(fiber.rxdbm2) if fiber.rxdbm2 else '')
            ws.cell(row=idx, column=14, value=float(fiber.dl_loss) if fiber.dl_loss else '')
            ws.cell(row=idx, column=15, value=float(fiber.ul_loss) if fiber.ul_loss else '')
            ws.cell(row=idx, column=16, value=fiber.length or '')
            ws.cell(row=idx, column=17, value='Sì' if fiber.is_dl_critical else 'No')
            ws.cell(row=idx, column=18, value='Sì' if fiber.is_ul_critical else 'No')
            ws.cell(row=idx, column=19, value='Sì' if fiber.is_link_down else 'No')
            
            if fiber.is_dl_critical or fiber.is_ul_critical or fiber.is_link_down:
                self._row_fill(ws, idx, BG_CRITICAL)
        
        self._autosize(ws)


    # ------------------------------------------------------------------ #
    #  8. SFP Modules                                                       #
    # ------------------------------------------------------------------ #

    def _sheet_sfp(self):
        ws = self.wb.create_sheet("SFP Modules")
        headers = ['Port', 'RiL', 'Board', 'LNH', 'Vendor', 'Device Name', 'REV', 
                   'Serial', 'Date', 'Ericsson Product', 'WL (nm)', 'Temp (°C)', 
                   'TXbs (%)', 'TX (dBm)', 'RX (dBm)', 'TN', 'RX Critical']
        self._header_style(ws, headers, COLOR_GOLD)
        
        for idx, sfp in enumerate(self.analysis.sfp_modules.all(), start=2):
            ws.cell(row=idx, column=1, value=sfp.port or '')
            ws.cell(row=idx, column=2, value=sfp.ril or '')
            ws.cell(row=idx, column=3, value=sfp.board or '')
            ws.cell(row=idx, column=4, value=sfp.lnh or '')
            ws.cell(row=idx, column=5, value=sfp.vendor or '')
            ws.cell(row=idx, column=6, value=sfp.device_name or '')
            ws.cell(row=idx, column=7, value=sfp.rev or '')
            ws.cell(row=idx, column=8, value=sfp.serial or '')
            ws.cell(row=idx, column=9, value=sfp.date or '')
            ws.cell(row=idx, column=10, value=sfp.ericsson_product or '')
            ws.cell(row=idx, column=11, value=float(sfp.wl) if sfp.wl else '')
            ws.cell(row=idx, column=12, value=sfp.temperature if sfp.temperature else '')
            ws.cell(row=idx, column=13, value=sfp.txbs if sfp.txbs else '')
            ws.cell(row=idx, column=14, value=float(sfp.tx_dbm) if sfp.tx_dbm else '')
            ws.cell(row=idx, column=15, value=float(sfp.rx_dbm) if sfp.rx_dbm else '')
            ws.cell(row=idx, column=16, value='Sì' if sfp.is_tn_backhaul else 'No')
            ws.cell(row=idx, column=17, value='Sì' if sfp.is_rx_critical else 'No')
            
            if sfp.is_rx_critical:
                self._row_fill(ws, idx, BG_CRITICAL)
            elif sfp.is_tn_backhaul:
                self._row_fill(ws, idx, "E3F2FD")
        
        self._autosize(ws)



    # ------------------------------------------------------------------ #
    #  9. Branch Pairs                                                     #
    # ------------------------------------------------------------------ #

    def _sheet_branch_pairs(self):
        ws = self.wb.create_sheet("Branch Pairs")
        headers = ['FRU', 'Board', 'RF Port', 'Branch Pair', 'Risultato',
                   'Warning (OKW)', 'Critical (NOK)']
        self._header_style(ws, headers, COLOR_DARK)

        for idx, bp in enumerate(self.analysis.branch_pairs.all(), start=2):
            ws.cell(row=idx, column=1, value=bp.fru)
            ws.cell(row=idx, column=2, value=bp.board)
            ws.cell(row=idx, column=3, value=bp.rf_port)
            ws.cell(row=idx, column=4, value=bp.branch_pair)
            ws.cell(row=idx, column=5, value=bp.result)
            ws.cell(row=idx, column=6, value='Sì' if bp.is_warning else 'No')
            ws.cell(row=idx, column=7, value='Sì' if bp.is_critical else 'No')
            if bp.is_critical:
                self._row_fill(ws, idx, BG_CRITICAL)
            elif bp.is_warning:
                self._row_fill(ws, idx, BG_WARNING)

        self._autosize(ws)

    # ------------------------------------------------------------------ #
    #  10. LGA Alarms                                                      #
    # ------------------------------------------------------------------ #
    def _sheet_lga(self):
        ws = self.wb.create_sheet("LGA Alarms")
        headers = [
            'Timestamp (UTC)', 'Severità', 'Codice Sev',
            'Specific Problem', 'Managed Object', 'Additional Info',
        ]
        self._header_style(ws, headers, COLOR_NAVY)

        # Mappa codice → etichetta leggibile
        SEV_LABEL = {
            'C': 'Critical',
            'M': 'Major',
            'm': 'minor',
            'w': 'Warning',
            '*': 'Ceasing',
        }
        # Colori riga per severità
        SEV_COLOR = {
            'C': BG_CRITICAL,   # rosso chiaro
            'M': BG_WARNING,    # giallo chiaro
        }

        for idx, alarm in enumerate(self.analysis.lga_alarms.all(), start=2):
            # Timestamp come stringa leggibile
            ts_str = alarm.timestamp.strftime('%Y-%m-%d %H:%M:%S') if alarm.timestamp else ''
            ws.cell(row=idx, column=1, value=ts_str)
            ws.cell(row=idx, column=2, value=SEV_LABEL.get(alarm.severity, alarm.severity))
            ws.cell(row=idx, column=3, value=alarm.severity)
            ws.cell(row=idx, column=4, value=alarm.specific_problem)
            ws.cell(row=idx, column=5, value=alarm.managed_object)
            ws.cell(row=idx, column=6, value=alarm.additional_info)

            # Evidenziazione per Critical e Major; grigio chiaro per Ceasing
            if alarm.severity in SEV_COLOR:
                self._row_fill(ws, idx, SEV_COLOR[alarm.severity])
            elif alarm.severity == '*':
                self._row_fill(ws, idx, "EFEFEF")

        self._autosize(ws)

